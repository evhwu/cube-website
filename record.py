import requests
from time import sleep
import json
from pathlib import Path
import openpyxl
from openpyxl import styles

input_path = Path.cwd().joinpath("input", "json", "draft_input.json")

def record_draft():
    with input_path.open('r', encoding="utf-8") as f:
        draft_input = json.loads(f.read())
        
    confirm = int(input("Check for misspellings? 1 - All, 2 - Ignore Short, 3 - Skip: "))
    if confirm != 3:
        all_cards = [card for player in draft_input["color_order"].values()
                     for card in player]
        for card in all_cards:
            request_string = f"https://api.scryfall.com/cards/search?q={card}"
            response = requests.get(request_string, params = {"format" : "json"})
            try:
                scryfall_dict = response.json()["data"]
                for scryfall_entry in scryfall_dict:
                    if card != scryfall_entry['name'] and confirm == 1:
                        print(f"{card} -- {scryfall_entry['name']}")
            except:
                print(card)
            sleep(.6)
    
    wb = openpyxl.Workbook()
    base_style = styles.NamedStyle(name="base",
                                   font=styles.Font(name="Cambria",
                                                    size=12),
                                                    alignment=styles.Alignment(horizontal="center"))
    
    def write_cell(sheet, row, col, value, color = None):
        cell = sheet.cell(row=row, column=col)
        cell.value = value
        if color:
            cell.fill = styles.PatternFill(start_color=color_order[color][1],
                                           fill_type="solid")
    
    draft_sheet = wb.active
    draft_sheet.title ="Draft"
    color_order = {"Green" : [0, "ff78d05c"], "Blue" : [1, "ff6FBBEA"],
                "Red" : [2, "ffEA6F6F"], "Purple" : [3, "ffAE6FEA"]}
    deck_sheet = wb.create_sheet(title = "Play")
    


record_draft()