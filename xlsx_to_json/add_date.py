from openpyxl import load_workbook
import os

BASE_PATH = os.getcwd()
save_path = os.path.dirname(f"{BASE_PATH}/archive/")

def copy_headings():
    archive_files = os.listdir(save_path)
    for f in archive_files:
        wb = load_workbook(os.path.join(save_path, f))
        current_sheet = wb["Draft"]
        
        current_sheet["F1"] = current_sheet["A1"].value
        current_sheet["G1"] = current_sheet["B1"].value
        current_sheet["H1"] = current_sheet["C1"].value
        current_sheet["I1"] = current_sheet["D1"].value
        wb.save(os.path.join(save_path, f))
def add_date():
    archive_files = os.listdir(save_path)

    patch = "9.0"
    for f in archive_files:
        file_num, file_ext = os.path.splitext(f)
        
        wb = load_workbook(os.path.join(save_path, f))
        wb.create_sheet(title="Date")
        current_sheet = wb["Date"]
        print(f"{file_num}: ")
        date = input("Insert Date: (ex. 11/20/2020): ")
        patch_input = input(f"Insert Patch Number (ex. 2.1, 3.0) or Enter for last value ({patch}):  ")
        if not patch_input == "":
            patch = patch_input

        current_sheet["A1"] = date
        current_sheet["A2"] = patch   
        wb.save(os.path.join(save_path, f))
    